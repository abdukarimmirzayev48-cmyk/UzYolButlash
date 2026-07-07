from __future__ import annotations

import re
import sys
import os
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select


PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
os.environ.setdefault("DATABASE_URL", f"sqlite:///{PROJECT_ROOT / 'bitum.db'}")

from backend.app.db.session import SessionLocal  # noqa: E402
from backend.app.models.customer_request import CompanyRegistry  # noqa: E402


EXCEL_PATH = PROJECT_ROOT / "backend" / "data" / "road_organizations.xlsx"


@dataclass
class ImportSummary:
    total_rows_read: int = 0
    organizations_imported: int = 0
    skipped_rows: int = 0
    duplicates_updated: int = 0


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def normalize_inn(value: Any) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+(\.0)?", text):
        return text.split(".")[0]
    digits = re.sub(r"\D", "", text)
    return digits if digits and digits == text else None


def utc_now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def header_key(value: Any) -> str:
    return (clean_text(value) or "").lower().replace("\n", " ")


def detect_header(row: tuple[Any, ...]) -> dict[str, int] | None:
    columns = {header_key(value): index for index, value in enumerate(row)}
    joined = " ".join(columns)
    if not (("стир" in joined or "stir" in joined) and ("корхона" in joined or "ташкилот" in joined)):
        return None

    def find(*needles: str) -> int | None:
        for key, index in columns.items():
            if all(needle in key for needle in needles):
                return index
        return None

    return {
        "inn": find("стир") or find("stir") or 2,
        "company_name": find("корхона") or find("ташкилот") or 3,
        "privatization_project_name": find("хусусийлаштириш") or 4,
        "legal_address": find("юридик") or find("manzil") or 5,
        "activity_type": find("асосий", "фаолият") or 6,
        "function_description": find("функция") or find("вазифа") or 7,
    }


def row_value(row: tuple[Any, ...], index: int | None) -> str | None:
    if index is None or index >= len(row):
        return None
    return clean_text(row[index])


def looks_like_region(row: tuple[Any, ...]) -> str | None:
    values = [clean_text(value) for value in row if clean_text(value)]
    if len(values) != 1:
        return None
    value = values[0] or ""
    lower = value.lower()
    if any(marker in lower for marker in ("вилояти", "viloyati", "шаҳри", "shahri", "республикаси")):
        return value
    return None


def import_company_registry(path: Path = EXCEL_PATH) -> ImportSummary:
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    summary = ImportSummary()
    header: dict[str, int] | None = None
    current_region: str | None = None
    seen_inns: set[str] = set()

    db = SessionLocal()
    try:
        for row in worksheet.iter_rows(values_only=True):
            summary.total_rows_read += 1
            if header is None:
                header = detect_header(row)
                if header is None:
                    summary.skipped_rows += 1
                continue

            region = looks_like_region(row)
            if region:
                current_region = region
                summary.skipped_rows += 1
                continue

            inn = normalize_inn(row[header["inn"]])
            company_name = row_value(row, header["company_name"])
            if not inn or not company_name:
                summary.skipped_rows += 1
                continue

            values = {
                "inn": inn,
                "company_name": company_name,
                "region": current_region,
                "legal_address": row_value(row, header["legal_address"]),
                "activity_type": row_value(row, header["activity_type"]),
                "function_description": row_value(row, header["function_description"]),
                "privatization_project_name": row_value(row, header["privatization_project_name"]),
                "updated_at": utc_now(),
            }

            existing = db.scalar(select(CompanyRegistry).where(CompanyRegistry.inn == inn))
            if existing:
                for key, value in values.items():
                    setattr(existing, key, value)
                summary.duplicates_updated += 1
            else:
                db.add(CompanyRegistry(**values, created_at=utc_now()))
                summary.organizations_imported += 1
            seen_inns.add(inn)

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
        workbook.close()
    return summary


def main() -> None:
    summary = import_company_registry()
    print("Company registry import completed:")
    print(f"  total rows read: {summary.total_rows_read}")
    print(f"  organizations imported: {summary.organizations_imported}")
    print(f"  skipped rows: {summary.skipped_rows}")
    print(f"  duplicates skipped/updated: {summary.duplicates_updated}")


if __name__ == "__main__":
    main()
