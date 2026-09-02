"""ABZ nuqtalari ro'yxatining Excel eksporti.

Ekranda nima ko'rinsa, o'sha eksport qilinadi: qatorlar ham, filtr ham API
qaytargan ro'yxatning o'zi. Yorliqlar server tomonida yoziladi, chunki
ikkilik faylni brauzerda o'girib bo'lmaydi.
"""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="176B5B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
TOTAL_FONT = Font(bold=True)
ATTENTION_FILL = PatternFill("solid", fgColor="FDF3E3")
INACTIVE_FILL = PatternFill("solid", fgColor="FBECEB")
THIN = Side(style="thin", color="D9E0E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LABELS = {
    "lat": {
        "title": "ABZ nuqtalari",
        "sheet": "ABZ",
        "station_title": "Temiryo'l stansiyalari",
        "station_sheet": "Stansiyalar",
        "name": "ABZ",
        "code": "Kodi",
        "client": "Mijoz",
        "region": "Viloyat",
        "district": "Tuman",
        "address": "Manzil",
        "capacity": "Quvvati, t/kun",
        "tank": "Sisterna, t",
        "responsible": "Mas'ul shaxs",
        "position": "Lavozimi",
        "phone": "Telefon",
        "hours": "Ish vaqti",
        "latitude": "Kenglik",
        "longitude": "Uzunlik",
        "status": "Holati",
        "updated": "Yangilangan",
        "total": "Jami",
        "active": "Faol",
        "attention": "E'tibor talab qiladi",
        "inactive": "Faol emas",
        "planned": "Rejalashtirilgan",
    },
    "cyr": {
        "title": "АБЗ нуқталари",
        "sheet": "АБЗ",
        "station_title": "Темирйўл стансиялари",
        "station_sheet": "Стансиялар",
        "name": "АБЗ",
        "code": "Коди",
        "client": "Мижоз",
        "region": "Вилоят",
        "district": "Туман",
        "address": "Манзил",
        "capacity": "Қуввати, т/кун",
        "tank": "Систерна, т",
        "responsible": "Масъул шахс",
        "position": "Лавозими",
        "phone": "Телефон",
        "hours": "Иш вақти",
        "latitude": "Кенглик",
        "longitude": "Узунлик",
        "status": "Ҳолати",
        "updated": "Янгиланган",
        "total": "Жами",
        "active": "Фаол",
        "attention": "Эътибор талаб қилади",
        "inactive": "Фаол эмас",
        "planned": "Режалаштирилган",
    },
}

COLUMNS = [
    ("name", 22),
    ("code", 12),
    ("client", 30),
    ("region", 22),
    ("district", 20),
    ("address", 34),
    ("capacity", 15),
    ("tank", 13),
    ("responsible", 24),
    ("position", 18),
    ("phone", 20),
    ("hours", 16),
    ("latitude", 12),
    ("longitude", 12),
    ("status", 20),
    ("updated", 18),
]


def _num(value) -> float | None:
    return None if value is None else float(Decimal(value))


def build_workbook(points: list, lang: str = "cyr", station: bool = False) -> BytesIO:
    words = LABELS.get(lang, LABELS["cyr"])
    book = Workbook()
    sheet = book.active
    sheet.title = words["station_sheet"] if station else words["sheet"]

    sheet["A1"] = words["station_title"] if station else words["title"]
    sheet["A1"].font = TITLE_FONT

    header_row = 3
    for index, (key, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=index, value=words[key])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width

    row_number = header_row
    capacity_total = Decimal("0")
    for point in points:
        row_number += 1
        capacity_total += Decimal(point.daily_capacity_tons or 0)
        status = point.status.value
        values = [
            point.name,
            point.code,
            point.client.name if point.client else None,
            point.region,
            point.district,
            point.address,
            _num(point.daily_capacity_tons),
            _num(point.tank_capacity_tons),
            point.responsible_name,
            point.responsible_position,
            point.responsible_phone,
            point.working_hours,
            point.latitude,
            point.longitude,
            words.get(status, status),
            point.updated_at.strftime("%Y-%m-%d %H:%M") if point.updated_at else None,
        ]
        # E'tibor talab qiladigan va yopilgan nuqtalar faylda ham ajralib
        # tursin: xulosa aynan shu qatorlar uchun ochiladi.
        fill = ATTENTION_FILL if status == "attention" else INACTIVE_FILL if status == "inactive" else None
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=index, value=value)
            cell.border = BORDER
            if fill:
                cell.fill = fill

    row_number += 1
    sheet.cell(row=row_number, column=1, value=words["total"]).font = TOTAL_FONT
    total_cell = sheet.cell(row=row_number, column=7, value=float(capacity_total))
    total_cell.font = TOTAL_FONT
    total_cell.border = BORDER

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(points)}"

    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream
