"""Park xulosasining Excel eksporti.

Bu bo'lim qog'ozdagi jadvalning o'rnini bosadi, lekin xulosani rahbariyatga
yuborish kerak bo'lganda fayl baribir talab qilinadi. Shuning uchun ekranda
nima ko'rinsa, o'sha eksport qilinadi: qatorlar ham, davr ham API
qaytargan xulosaning o'zi -- fayl bilan ekran filtr nimani tanlagani
haqida bahslasha olmaydi.

Yorliqlar server tomonida yoziladi: ikkilik faylni brauzerda o'girib
bo'lmaydi, shuning uchun ikkala alifbo ham shu yerda turadi.
"""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="176B5B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
TOTAL_FONT = Font(bold=True)
RISK_FILL = PatternFill("solid", fgColor="FBE9E7")
THIN = Side(style="thin", color="D9E0E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LABELS = {
    "lat": {
        "title": "Park bo'yicha xulosa",
        "period": "Davr",
        "all_time": "Butun davr",
        "sheet": "Xulosa",
        "vehicle": "Mashina",
        "driver": "Haydovchi",
        "status": "Holati",
        "trips": "Reyslar",
        "tons": "Tashilgan, t",
        "distance": "Masofa, km",
        "fuel": "Yoqilg'i, l",
        "norm": "Norma, l",
        "difference": "Chetlanish, l",
        "difference_percent": "Chetlanish, %",
        "suspected": "Slivga shubha, l",
        "events": "Hodisalar",
        "unchecked": "Tekshirilmagan",
        "damage": "Undirilgan zarar",
        "downtime": "Ta'mir turib qolishi, soat",
        "repair_amount": "Ta'mir xarajati",
        "to_service": "TO gacha, km",
        "documents": "Hujjatlar",
        "total": "Jami",
        "ok": "Joyida",
        "soon": "Muddati tugayapti",
        "expired": "Muddati o'tgan",
        "unknown": "Kiritilmagan",
        "free": "Bo'sh",
        "repair": "Ta'mirda",
        "service": "Texnik xizmatda",
        "idle": "Bekor turibdi",
        "inactive": "Parkda emas",
    },
    "cyr": {
        "title": "Парк бўйича хулоса",
        "period": "Давр",
        "all_time": "Бутун давр",
        "sheet": "Хулоса",
        "vehicle": "Машина",
        "driver": "Ҳайдовчи",
        "status": "Ҳолати",
        "trips": "Рейслар",
        "tons": "Ташилган, т",
        "distance": "Масофа, км",
        "fuel": "Ёқилғи, л",
        "norm": "Норма, л",
        "difference": "Четланиш, л",
        "difference_percent": "Четланиш, %",
        "suspected": "Сливга шубҳа, л",
        "events": "Ҳодисалар",
        "unchecked": "Текширилмаган",
        "damage": "Ундирилган зарар",
        "downtime": "Таъмир туриб қолиши, соат",
        "repair_amount": "Таъмир харажати",
        "to_service": "ТО гача, км",
        "documents": "Ҳужжатлар",
        "total": "Жами",
        "ok": "Жойида",
        "soon": "Муддати тугаяпти",
        "expired": "Муддати ўтган",
        "unknown": "Киритилмаган",
        "free": "Бўш",
        "repair": "Таъмирда",
        "service": "Техник хизматда",
        "idle": "Бекор турибди",
        "inactive": "Паркда эмас",
    },
}

COLUMNS = [
    ("vehicle", 18),
    ("driver", 26),
    ("status", 16),
    ("trips", 10),
    ("tons", 13),
    ("distance", 13),
    ("fuel", 12),
    ("norm", 12),
    ("difference", 14),
    ("difference_percent", 14),
    ("suspected", 16),
    ("events", 12),
    ("unchecked", 15),
    ("damage", 18),
    ("downtime", 22),
    ("repair_amount", 18),
    ("to_service", 14),
    ("documents", 18),
]


def _num(value) -> float | None:
    if value is None:
        return None
    return float(Decimal(value))


def build_workbook(summary, lang: str = "cyr") -> BytesIO:
    words = LABELS.get(lang, LABELS["cyr"])
    book = Workbook()
    sheet = book.active
    sheet.title = words["sheet"]

    sheet["A1"] = words["title"]
    sheet["A1"].font = TITLE_FONT
    period = words["all_time"]
    if summary.date_from or summary.date_to:
        period = f"{summary.date_from or '…'} — {summary.date_to or '…'}"
    sheet["A2"] = f"{words['period']}: {period}"

    header_row = 4
    for index, (key, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=index, value=words[key])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width

    row_number = header_row
    for row in summary.rows:
        row_number += 1
        values = [
            row.vehicle_number,
            row.driver_name,
            words.get(row.status, row.status),
            row.trip_count,
            _num(row.delivered_tons),
            _num(row.distance_km),
            _num(row.fuel_liters),
            _num(row.norm_liters),
            _num(row.difference_liters),
            _num(row.difference_percent),
            _num(row.suspected_liters),
            row.event_count,
            row.unchecked_event_count,
            _num(row.damage_amount),
            _num(row.repair_downtime_hours),
            _num(row.repair_amount),
            _num(row.remaining_to_service_km),
            words.get(row.document_level, row.document_level),
        ]
        # Slivga shubhasi bor yoki hujjati o'tgan mashina ko'rinib tursin:
        # xulosa aynan shu qatorlar uchun ochiladi.
        risky = (row.suspected_liters or 0) > 0 or row.document_level == "expired"
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=index, value=value)
            cell.border = BORDER
            if risky:
                cell.fill = RISK_FILL

    totals = summary.totals
    row_number += 1
    total_values = [
        words["total"],
        None,
        None,
        totals.trip_count,
        _num(totals.delivered_tons),
        _num(totals.distance_km),
        _num(totals.fuel_liters),
        _num(totals.norm_liters),
        _num(totals.difference_liters),
        None,
        _num(totals.suspected_liters),
        totals.event_count,
        totals.unchecked_event_count,
        _num(totals.damage_amount),
        _num(totals.repair_downtime_hours),
        _num(totals.repair_amount),
        None,
        None,
    ]
    for index, value in enumerate(total_values, start=1):
        cell = sheet.cell(row=row_number, column=index, value=value)
        cell.font = TOTAL_FONT
        cell.border = BORDER

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(summary.rows)}"

    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream
