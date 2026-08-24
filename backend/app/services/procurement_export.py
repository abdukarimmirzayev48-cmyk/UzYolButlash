"""Xaridlar ro'yxatining Excel eksporti.

client_export bilan bir xil shakl: yorliqlar server tomonida yoziladi, chunki
ikkilik faylni brauzerda o'girib bo'lmaydi -- ikkala alifbo ham shu yerda
turadi va chaqiruvchi o'quvchining tilini uzatadi.

Qatorlar ekranda chizilayotgan ProcurementListItem obyektlarining o'zi, shuning
uchun fayl bilan ekran filtr nimani tanlagani haqida hech qachon bahslasha
olmaydi.
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="176B5B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="D9E0E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_LABELS = {
    "lat": {
        "draft": "Qoralama",
        "supplier_search": "Ta'minotchi qidirilmoqda",
        "offers_received": "Takliflar olindi",
        "supplier_selected": "Ta'minotchi tanlandi",
        "supplier_confirmed": "Ta'minotchi tasdiqlandi",
        "purchase_approved": "Xarid tasdiqlandi",
        "waiting_supplier_ready": "Ta'minotchi tayyorlanmoqda",
        "ready_for_pickup": "Olib ketishga tayyor",
        "ready_for_delivery": "Yetkazishga tayyor",
        "completed": "Yakunlandi",
        "cancelled": "Bekor qilindi",
        "issue": "Muammo",
    },
    "cyr": {
        "draft": "Қоралама",
        "supplier_search": "Таъминотчи қидирилмоқда",
        "offers_received": "Таклифлар олинди",
        "supplier_selected": "Таъминотчи танланди",
        "supplier_confirmed": "Таъминотчи тасдиқланди",
        "purchase_approved": "Харид тасдиқланди",
        "waiting_supplier_ready": "Таъминотчи тайёрланмоқда",
        "ready_for_pickup": "Олиб кетишга тайёр",
        "ready_for_delivery": "Етказишга тайёр",
        "completed": "Якунланди",
        "cancelled": "Бекор қилинди",
        "issue": "Муаммо",
    },
}

LABELS = {
    "lat": {
        "sheet": "Xaridlar",
        "title": "Xaridlar ro'yxati",
        "columns": [
            ("#", 6),
            ("Xarid raqami", 30),
            ("Sana", 12),
            ("Buyurtma", 26),
            ("Mijoz", 44),
            ("Shartnoma", 20),
            ("Mahsulot", 32),
            ("Birlik", 10),
            ("Miqdor", 14),
            ("Tanlangan miqdor", 18),
            ("Takliflar", 12),
            ("Tanlangan ta'minotchilar", 22),
            ("Summa", 20),
            ("Status", 24),
        ],
        "generated": "Yuklab olingan sana",
        "filters": "Tanlangan filtrlar",
        "no_filters": "Filtrsiz (barcha xaridlar)",
        "total_row": "Jami",
        "count": "Xaridlar soni",
    },
    "cyr": {
        "sheet": "Харидлар",
        "title": "Харидлар рўйхати",
        "columns": [
            ("#", 6),
            ("Харид рақами", 30),
            ("Сана", 12),
            ("Буюртма", 26),
            ("Мижоз", 44),
            ("Шартнома", 20),
            ("Маҳсулот", 32),
            ("Бирлик", 10),
            ("Миқдор", 14),
            ("Танланган миқдор", 18),
            ("Таклифлар", 12),
            ("Танланган таъминотчилар", 22),
            ("Сумма", 20),
            ("Статус", 24),
        ],
        "generated": "Юклаб олинган сана",
        "filters": "Танланган филтрлар",
        "no_filters": "Филтрсиз (барча харидлар)",
        "total_row": "Жами",
        "count": "Харидлар сони",
    },
}

# Uzun matnli ustunlar -- ular o'ralib chiqadi.
WRAP_COLUMNS = {5, 7}
MONEY_COLUMN = 13
QUANTITY_COLUMNS = (9, 10)


def labels(lang: str) -> dict:
    return LABELS.get(lang, LABELS["cyr"])


def status_label(value, lang: str) -> str:
    key = getattr(value, "value", value)
    return STATUS_LABELS.get(lang, STATUS_LABELS["cyr"]).get(key, key or "")


def cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, Decimal):
        return float(value)
    return value


def build_workbook(items: list, lang: str, filter_note: str = "") -> BytesIO:
    text = labels(lang)
    columns = text["columns"]

    book = Workbook()
    sheet = book.active
    sheet.title = text["sheet"]

    sheet.cell(row=1, column=1, value=text["title"]).font = TITLE_FONT
    sheet.cell(row=2, column=1, value=f"{text['generated']}: {date.today():%d.%m.%Y}")
    sheet.cell(row=3, column=1, value=f"{text['filters']}: {filter_note or text['no_filters']}")
    sheet.cell(row=4, column=1, value=f"{text['count']}: {len(items)}")

    header_row = 6
    for index, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width

    for offset, item in enumerate(items):
        values = [
            offset + 1,
            item.procurement_number,
            item.procurement_date,
            item.order.order_number if item.order else None,
            item.client.name if item.client else None,
            item.contract.contract_number if item.contract else None,
            item.product,
            item.unit,
            item.required_quantity,
            item.selected_quantity,
            item.offers_count,
            item.selected_suppliers_count,
            item.purchase_amount,
            status_label(item.status, lang),
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=header_row + 1 + offset, column=index, value=cell_value(value))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=index in WRAP_COLUMNS)
            if index == 3 and isinstance(value, date):
                cell.number_format = "DD.MM.YYYY"
            if index == MONEY_COLUMN:
                cell.number_format = "# ##0"
            if index in QUANTITY_COLUMNS:
                cell.number_format = "# ##0.###"

    if items:
        total_row = header_row + 1 + len(items)
        sheet.cell(row=total_row, column=1, value=text["total_row"]).font = Font(bold=True)
        for index, getter in ((9, "required_quantity"), (10, "selected_quantity"), (MONEY_COLUMN, "purchase_amount")):
            cell = sheet.cell(row=total_row, column=index)
            cell.value = float(sum((getattr(item, getter) or Decimal("0")) for item in items))
            cell.font = Font(bold=True)
            cell.border = BORDER
            cell.number_format = "# ##0" if index == MONEY_COLUMN else "# ##0.###"

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(items)}"

    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream
