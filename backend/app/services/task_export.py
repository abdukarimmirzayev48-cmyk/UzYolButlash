"""Excel export for Ijro.

Labels are written server-side, so unlike the HTML UI they can't be
transliterated in the browser -- both alphabets are spelled out here and the
caller passes the reader's current language.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.app.models.task import Task
from backend.app.services import task_stats

HEADER_FILL = PatternFill("solid", fgColor="176B5B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="D9E0E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OVERDUE_FILL = PatternFill("solid", fgColor="FDE7E4")

LABELS = {
    "lat": {
        "sheet_tasks": "Topshiriqlar",
        "sheet_summary": "Xulosa",
        "columns": [
            ("#", 6),
            ("Sarlavha", 44),
            ("Holat", 16),
            ("Muhimlik", 12),
            ("Bo'lim", 24),
            ("Mas'ul xodimlar", 34),
            ("Kim tomonidan berilgan", 22),
            ("Yaratuvchi", 22),
            ("Yaratilgan sana", 16),
            ("Muddat", 14),
            ("Qolgan kun", 12),
            ("Muddati o'tgan", 14),
            ("Bajarilgan sana", 16),
            ("Yopilgan sana", 16),
            ("O'z vaqtida", 12),
            ("Izohlar", 9),
            ("Fayllar", 9),
            ("Tavsif", 50),
        ],
        "status": {
            "new": "Yangi",
            "accepted": "Qabul qilindi",
            "in_progress": "Bajarilmoqda",
            "done": "Bajarildi",
            "verified": "Tasdiqlandi",
            "rejected": "Rad etildi",
        },
        "priority": {"low": "Past", "medium": "O'rta", "high": "Yuqori", "urgent": "Shoshilinch"},
        "yes": "Ha",
        "no": "Yo'q",
        "generated": "Yuklab olingan sana",
        "filters": "Tanlangan filtrlar",
        "total_row": "Jami",
        "by_status": "Holat bo'yicha",
        "by_priority": "Muhimlik bo'yicha",
        "by_department": "Bo'lim bo'yicha",
        "by_employee": "Xodim bo'yicha",
        "count": "Soni",
        "open": "Ochiq",
        "overdue": "Muddati o'tgan",
        "completed": "Bajarilgan",
        "on_time": "O'z vaqtida",
        "name": "Nomi",
        "no_filters": "Filtrsiz (barcha topshiriqlar)",
    },
    "cyr": {
        "sheet_tasks": "Топшириқлар",
        "sheet_summary": "Хулоса",
        "columns": [
            ("#", 6),
            ("Сарлавҳа", 44),
            ("Ҳолат", 16),
            ("Муҳимлик", 12),
            ("Бўлим", 24),
            ("Масъул ходимлар", 34),
            ("Ким томонидан берилган", 22),
            ("Яратувчи", 22),
            ("Яратилган сана", 16),
            ("Муддат", 14),
            ("Қолган кун", 12),
            ("Муддати ўтган", 14),
            ("Бажарилган сана", 16),
            ("Ёпилган сана", 16),
            ("Ўз вақтида", 12),
            ("Изоҳлар", 9),
            ("Файллар", 9),
            ("Тавсиф", 50),
        ],
        "status": {
            "new": "Янги",
            "accepted": "Қабул қилинди",
            "in_progress": "Бажарилмоқда",
            "done": "Бажарилди",
            "verified": "Тасдиқланди",
            "rejected": "Рад этилди",
        },
        "priority": {"low": "Паст", "medium": "Ўрта", "high": "Юқори", "urgent": "Шошилинч"},
        "yes": "Ҳа",
        "no": "Йўқ",
        "generated": "Юклаб олинган сана",
        "filters": "Танланган филтрлар",
        "total_row": "Жами",
        "by_status": "Ҳолат бўйича",
        "by_priority": "Муҳимлик бўйича",
        "by_department": "Бўлим бўйича",
        "by_employee": "Ходим бўйича",
        "count": "Сони",
        "open": "Очиқ",
        "overdue": "Муддати ўтган",
        "completed": "Бажарилган",
        "on_time": "Ўз вақтида",
        "name": "Номи",
        "no_filters": "Филтрсиз (барча топшириқлар)",
    },
}


def labels_for(lang: str | None) -> dict:
    return LABELS.get((lang or "cyr").lower(), LABELS["cyr"])


def enum_value(value) -> str:
    return value.value if hasattr(value, "value") else str(value or "")


def write_header(sheet, columns: list[tuple[str, int]], row: int = 1) -> None:
    for index, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=row, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[row].height = 28


def build_tasks_sheet(sheet, tasks: list[Task], labels: dict, now: datetime) -> None:
    columns = labels["columns"]
    write_header(sheet, columns)

    for offset, task in enumerate(tasks):
        row = offset + 2
        overdue = task_stats.is_overdue(task, now)
        on_time = task_stats.completed_on_time(task)
        days_left = (task.deadline.date() - now.date()).days if task.deadline else None
        values = [
            task.id,
            task.title,
            labels["status"].get(enum_value(task.status), enum_value(task.status)),
            labels["priority"].get(enum_value(task.priority), enum_value(task.priority)),
            task.department.name if task.department else "",
            task_stats.assignee_names(task),
            task.created_by or "",
            task.created_by_user.full_name if task.created_by_user else "",
            task.created_at,
            task.deadline,
            days_left,
            labels["yes"] if overdue else labels["no"],
            task.completed_at,
            task.closed_at,
            "" if on_time is None else (labels["yes"] if on_time else labels["no"]),
            len(task.comments),
            len(task.attachments),
            task.description or "",
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=index, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=index in (2, 6, 18))
            if isinstance(value, datetime):
                cell.number_format = "DD.MM.YYYY HH:MM" if index == 9 else "DD.MM.YYYY"
            if overdue:
                cell.fill = OVERDUE_FILL

    sheet.freeze_panes = "A2"
    if tasks:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(tasks) + 1}"


def write_block(sheet, row: int, title: str, headers: list[str], rows: list[list], widths: list[int]) -> int:
    cell = sheet.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    row += 1
    for index, header in enumerate(headers, start=1):
        head = sheet.cell(row=row, column=index, value=header)
        head.fill = HEADER_FILL
        head.font = HEADER_FONT
        head.border = BORDER
        head.alignment = Alignment(horizontal="center")
        current = sheet.column_dimensions[get_column_letter(index)].width or 0
        sheet.column_dimensions[get_column_letter(index)].width = max(current, widths[index - 1])
    row += 1
    for values in rows:
        for index, value in enumerate(values, start=1):
            body = sheet.cell(row=row, column=index, value=value)
            body.border = BORDER
        row += 1
    return row + 1


def build_summary_sheet(sheet, dashboard: dict, labels: dict, filter_note: str, now: datetime) -> None:
    sheet.cell(row=1, column=1, value=labels["sheet_summary"]).font = Font(bold=True, size=15)
    sheet.cell(row=2, column=1, value=labels["generated"])
    stamp = sheet.cell(row=2, column=2, value=now)
    stamp.number_format = "DD.MM.YYYY HH:MM"
    sheet.cell(row=3, column=1, value=labels["filters"])
    sheet.cell(row=3, column=2, value=filter_note or labels["no_filters"])

    summary = dashboard["summary"]
    row = write_block(
        sheet,
        5,
        labels["total_row"],
        [labels["name"], labels["count"]],
        [
            [labels["total_row"], summary["total"]],
            [labels["open"], summary["open"]],
            [labels["overdue"], summary["overdue"]],
            [labels["completed"], summary["completed"]],
            [labels["on_time"] + " %", summary["on_time_rate"]],
        ],
        [30, 14],
    )
    row = write_block(
        sheet,
        row,
        labels["by_status"],
        [labels["name"], labels["count"]],
        [[labels["status"].get(item["status"], item["status"]), item["count"]] for item in dashboard["by_status"]],
        [30, 14],
    )
    row = write_block(
        sheet,
        row,
        labels["by_priority"],
        [labels["name"], labels["count"]],
        [[labels["priority"].get(item["priority"], item["priority"]), item["count"]] for item in dashboard["by_priority"]],
        [30, 14],
    )
    row = write_block(
        sheet,
        row,
        labels["by_department"],
        [labels["name"], labels["count"], labels["open"], labels["overdue"], labels["completed"]],
        [
            [item["department"], item["total"], item["open"], item["overdue"], item["completed"]]
            for item in dashboard["by_department"]
        ],
        [30, 14, 12, 16, 14],
    )
    write_block(
        sheet,
        row,
        labels["by_employee"],
        [labels["name"], labels["count"], labels["open"], labels["overdue"], labels["completed"], labels["on_time"]],
        [
            [item["full_name"], item["total"], item["open"], item["overdue"], item["completed"], item["on_time"]]
            for item in dashboard["by_employee"]
        ],
        [30, 14, 12, 16, 14, 14],
    )


def build_workbook(tasks: list[Task], dashboard: dict, lang: str | None, filter_note: str = "") -> BytesIO:
    labels = labels_for(lang)
    now = datetime.now()
    workbook = Workbook()
    tasks_sheet = workbook.active
    tasks_sheet.title = labels["sheet_tasks"]
    build_tasks_sheet(tasks_sheet, tasks, labels, now)
    build_summary_sheet(workbook.create_sheet(labels["sheet_summary"]), dashboard, labels, filter_note, now)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
