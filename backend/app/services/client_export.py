"""Excel export for the client registry.

Same shape as task_export: the labels are written server-side, so unlike the
HTML UI they cannot be transliterated in the browser -- both alphabets are
spelled out here and the caller passes the reader's current language.

The rows are the same ClientListItem objects the table renders, so the file and
the screen can never disagree about what a filter selected.
"""

from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="176B5B")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="D9E0E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LABELS = {
    "lat": {
        "sheet": "Mijozlar",
        "title": "Mijozlar reyestri",
        "columns": [
            ("#", 6),
            ("Mijoz nomi", 46),
            ("STIR", 14),
            ("OKED", 10),
            ("Telefon", 18),
            ("Email", 26),
            ("Hudud", 24),
            ("Yuridik manzil", 46),
            ("Asosiy kontakt", 28),
            ("Kontakt lavozimi", 22),
            ("Kontakt telefoni", 18),
            ("Faol shartnomalar", 18),
            ("Faol buyurtmalar", 18),
            ("Qo'shilgan sana", 16),
        ],
        "generated": "Yuklab olingan sana",
        "filters": "Tanlangan filtrlar",
        "no_filters": "Filtrsiz (barcha mijozlar)",
        "total_row": "Jami",
        "count": "Mijozlar soni",
    },
    "cyr": {
        "sheet": "Мижозлар",
        "title": "Мижозлар реестри",
        "columns": [
            ("#", 6),
            ("Мижоз номи", 46),
            ("СТИР", 14),
            ("ОКЭД", 10),
            ("Телефон", 18),
            ("Э-почта", 26),
            ("Ҳудуд", 24),
            ("Юридик манзил", 46),
            ("Асосий контакт", 28),
            ("Контакт лавозими", 22),
            ("Контакт телефони", 18),
            ("Фаол шартномалар", 18),
            ("Фаол буюртмалар", 18),
            ("Қўшилган сана", 16),
        ],
        "generated": "Юклаб олинган сана",
        "filters": "Танланган филтрлар",
        "no_filters": "Филтрсиз (барча мижозлар)",
        "total_row": "Жами",
        "count": "Мижозлар сони",
    },
}


def labels(lang: str) -> dict:
    return LABELS.get(lang, LABELS["cyr"])


def cell_value(value):
    """Excel has no notion of our None, and a naive datetime writes cleanest."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def build_workbook(items: list, lang: str, filter_note: str = "") -> BytesIO:
    text = labels(lang)
    columns = text["columns"]

    book = Workbook()
    sheet = book.active
    sheet.title = text["sheet"]

    sheet.cell(row=1, column=1, value=text["title"]).font = TITLE_FONT
    sheet.cell(row=2, column=1, value=f"{text['generated']}: {date.today():%d.%m.%Y}")
    sheet.cell(row=3, column=1, value=f"{text['filters']}: {filter_note or text['no_filters']}")
    sheet.cell(row=4, column=1, value=f"{text['count']}: {len(items)}")

    header_row = 6
    for index, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = width

    for offset, item in enumerate(items):
        contact = item.primary_contact
        values = [
            offset + 1,
            item.name,
            item.inn,
            item.oked,
            item.phone,
            item.email,
            item.primary_region,
            item.legal_address,
            contact.full_name if contact else None,
            contact.position if contact else None,
            contact.phone if contact else None,
            item.active_contracts,
            item.active_orders,
            item.created_at,
        ]
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=header_row + 1 + offset, column=index, value=cell_value(value))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=index in {2, 8})
            if index == 14 and isinstance(value, datetime):
                cell.number_format = "DD.MM.YYYY"

    if items:
        total_row = header_row + 1 + len(items)
        sheet.cell(row=total_row, column=1, value=text["total_row"]).font = Font(bold=True)
        for index in (12, 13):
            cell = sheet.cell(row=total_row, column=index)
            cell.value = sum(
                (item.active_contracts if index == 12 else item.active_orders) or 0 for item in items
            )
            cell.font = Font(bold=True)
            cell.border = BORDER

    # Freeze under the header so the column names stay put while scrolling, and
    # turn on autofilter -- an exported registry is usually re-sorted in Excel.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{header_row + len(items)}"

    stream = BytesIO()
    book.save(stream)
    stream.seek(0)
    return stream
